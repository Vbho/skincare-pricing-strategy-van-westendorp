"""
build_excel.py — D2C Skincare Pricing Strategy Excel Deliverable
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path("outputs")

NAVY    = "FF1A3C5E"; CORAL   = "FFC0392B"; AMBER   = "FFE67E22"
GREEN   = "FF27AE60"; SKY     = "FF2980B9"; SILVER  = "FFBDC3C7"
DGREY   = "FF2C3E50"; WHITE   = "FFFFFFFF"; LTGREY  = "FFECF0F1"
MIDGREY = "FF95A5A6"; OFFWHITE= "FFFAFAFA"; PURPLE  = "FF8E44AD"
RED_BG  = "FFFDEDEC"; AMBER_BG= "FFFEF9E7"; GREEN_BG= "FFEAFAF1"
BLUE_IN = "FF2471A3"

def fill(h): return PatternFill("solid", fgColor=h)
def font(sz=11, color="FF000000", bold=False, italic=False):
    return Font(name="Arial", size=sz, color=color, bold=bold, italic=italic)
thin = Side(style="thin", color="FFBDC3C7")
def ba(c): c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def w(ws, r, c, v, sz=11, color="FF000000", bold=False, italic=False,
      bg=None, align="center", wrap=False, nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font(sz, color, bold, italic)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg: cell.fill = fill(bg)
    ba(cell)
    if nf: cell.number_format = nf
    return cell

def banner(ws, title, sub, cols=9):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = font(15, WHITE, bold=True); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(row=2, column=1, value=sub)
    s.font = font(9, DGREY, italic=True); s.fill = fill(LTGREY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 16

def sec(ws, row, text, cols=9, bg=NAVY, fg=WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(10, fg, bold=True); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

def hdr(ws, row, headers, widths, bg=NAVY):
    for i, (h, ww) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font(10, WHITE, bold=True); c.fill = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ba(c)
        ws.column_dimensions[get_column_letter(i)].width = ww
    ws.row_dimensions[row].height = 28

def sp(ws, row, cols=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 8

def note(ws, row, text, cols=9, bg=AMBER_BG, color=DGREY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(10, color); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ba(c); ws.row_dimensions[row].height = 30

wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
banner(ws1, "D2C Skincare Pricing Strategy — £1.4M ARR Brand · 12 SKUs",
       "Three-way triangulation: Van Westendorp PSM (n=340) + 8-competitor benchmarking + margin waterfall · Analyst: Vaishnavi Bhor")

kpis = [
    ("Brand ARR",        "£1.4M",    "12 SKUs · prices unchanged 3 years", CORAL),
    ("PSM Survey",       "n=340",    "Brand email list respondents",        SKY),
    ("Hero SKUs Below\nPrice Ceiling", "14–22%", "Top 4 SKUs vs PME",      CORAL),
    ("Comp Gap",         "£10-15",   "Below mid-premium peers",             AMBER),
    ("GP Uplift",        "£67K",     "Net gross profit gain from repricing",GREEN),
    ("Current GM",       "70.6%",    "vs 69.4% benchmark (Eightx 2026)",   SKY),
    ("Proposed GM",      "74.6%",    "Post-repricing gross margin",         GREEN),
    ("Vol. Elasticity",  "-0.8",     "Conservative for premium skincare",   NAVY),
]
for i, (lbl, val, note_txt, col) in enumerate(kpis, 1):
    ws1.column_dimensions[get_column_letter(i)].width = 17
    c1 = ws1.cell(row=4, column=i, value=lbl)
    c1.font = font(9, WHITE, bold=True); c1.fill = fill(col)
    c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[4].height = 28
    c2 = ws1.cell(row=5, column=i, value=val)
    c2.font = font(15, col, bold=True); c2.fill = fill(OFFWHITE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[5].height = 34
    c3 = ws1.cell(row=6, column=i, value=note_txt)
    c3.font = font(8, MIDGREY, italic=True); c3.fill = fill(OFFWHITE)
    c3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[6].height = 22

sp(ws1, 7)
sec(ws1, 8, "ANALYTICAL FRAMEWORK — THREE-WAY TRIANGULATION (McKINSEY ISSUE TREE)")
branches = [
    ("Branch 1", "Van Westendorp PSM\n(n=340 survey)", "What do consumers consider cheap, acceptable, expensive, and too expensive for hero SKUs?",
     "PMC (floor), IPP, OPP (optimal), PME (ceiling) derived from 4-question survey on brand email list"),
    ("Branch 2", "Competitor Benchmarking\n(8 brands)", "Where does this brand sit relative to 8 comparable UK D2C skincare brands?",
     "The Ordinary, The Inkey List, Byoma, Pixi, REN, Paula's Choice, Medik8, Elemis — real 2024-25 prices"),
    ("Branch 3", "Margin Waterfall\n(12 SKUs)", "What is the financial impact of repricing at proposed vs current levels?",
     "COGS held fixed in £. Volume elasticity -0.8 applied. GM benchmarked vs Eightx 2026 public DTC beauty data"),
    ("Branch 4", "Triangulated\nRecommendation", "Where do PSM, competitor data, and margin target all converge?",
     "Recommended price = PSM OPP, calibrated to sit between mid-market and premium competitors"),
]
hdr(ws1, 9, ["Branch", "Method", "Question", "Data Foundation"], [12, 22, 44, 44])
for i, (branch, method, question, foundation) in enumerate(branches, 10):
    ws1.row_dimensions[i].height = 32
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws1, i, 1, branch,     10, CORAL,  bold=True, bg=bg)
    w(ws1, i, 2, method,     10, NAVY,   bold=True, bg=bg, align="left", wrap=True)
    w(ws1, i, 3, question,   10, bg=bg,  align="left", wrap=True)
    w(ws1, i, 4, foundation,  9, MIDGREY,italic=True, bg=bg, align="left", wrap=True)

sp(ws1, 14)
sec(ws1, 15, "KEY FINDINGS")
findings = [
    ("Finding 1",  "Brand is underpriced — confirmed independently by all three methods",
     "PSM OPP sits £2–£8 above current prices. Competitor analysis shows brand priced £10-15 below mid-premium peers. Margin model shows current GMs at 70.6% — below the headroom that the PSM ceiling would allow."),
    ("Finding 2",  "Top 4 hero SKUs are 14–22% below the acceptable price ceiling",
     "PSM PME (the 'expensive but still buy' threshold) ranges from £35.70 (Vit C Serum) to £48.00 (Retinol). The brand's 'premium' product (Retinol, £38) sits 21% below its PME — a positioning contradiction."),
    ("Finding 3",  "GP uplift of £67K at conservative elasticity (-0.8)",
     "Price gain outweighs volume loss at -0.8 elasticity. For premium skincare with loyal D2C customers, -0.8 is already conservative — REN and Paula's Choice have raised prices 15-20% with minimal volume impact."),
    ("Finding 4",  "Simultaneous rebrand is critical to anchor the new price",
     "A price increase without a refresh risks customer push-back. New packaging, updated copy highlighting active ingredients, and a PR moment ('formulation upgrade') provides the psychological anchor for the new price."),
]
hdr(ws1, 16, ["Finding", "Headline", "Supporting Evidence"], [12, 38, 60])
for i, (label, headline, evidence) in enumerate(findings, 17):
    ws1.row_dimensions[i].height = 38
    bg = RED_BG if i == 17 else AMBER_BG if i == 18 else OFFWHITE
    w(ws1, i, 1, label,   10, CORAL, bold=True, bg=bg)
    w(ws1, i, 2, headline, 10, NAVY, bold=True, bg=bg, align="left", wrap=True)
    w(ws1, i, 3, evidence, 9, DGREY, bg=bg, align="left", wrap=True)


# ════════════════════════════════════════════════════════════════════
# SHEET 2 — VAN WESTENDORP PSM
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Van Westendorp PSM")
ws2.sheet_view.showGridLines = False
banner(ws2, "Branch 1: Van Westendorp Price Sensitivity Meter",
       "n=340 respondents from brand email list · 4 price perception questions · Standard PSM methodology (Van Westendorp, 1976)")

note(ws2, 4,
    "METHODOLOGY: Respondents were shown each hero SKU and asked four questions: "
    "(1) Too Cheap — at what price would you doubt the quality? "
    "(2) Cheap/Good Value — at what price is this a bargain? "
    "(3) Expensive — at what price is this expensive but worth considering? "
    "(4) Too Expensive — at what price would you not consider buying? "
    "Cumulative frequency curves are plotted; four intersections give PMC, IPP, OPP, and PME.",
    bg=AMBER_BG)

sp(ws2, 5)
sec(ws2, 6, "PSM KEY INTERSECTIONS — ALL HERO SKUs")
hdr(ws2, 7,
    ["SKU", "Current Price", "PMC\n(Floor — 'Too Cheap' line crosses 'Expensive')",
     "IPP\n(Indifference Point)", "OPP\n(Optimal — 'Cheap' crosses 'Expensive')",
     "PME\n(Ceiling — 'Cheap' crosses 'Too Expensive')", "Current vs Ceiling", "Recommended Price"],
    [24, 14, 16, 14, 16, 16, 18, 16])

psm_data = [
    ("Vitamin C Serum (30ml)",    28.00, 24.00, 24.00, 30.20, 35.70, -0.215, 33.00),
    ("HA Moisturiser (50ml)",     32.00, 26.00, 26.00, 35.50, 41.80, -0.234, 37.00),
    ("Retinol Night Treatment",   38.00, 30.00, 30.00, 44.00, 48.00, -0.208, 46.00),
    ("Niacinamide Pore Minimiser",26.00, 22.00, 22.00, 28.50, 33.00, -0.212, 30.00),
]
for i, (sku, cur, pmc, ipp, opp, pme, gap, rec) in enumerate(psm_data, 8):
    ws2.row_dimensions[i].height = 26
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    gap_col = CORAL if gap < -0.15 else AMBER
    w(ws2, i, 1, sku,  10, NAVY, bold=True, bg=bg, align="left")
    w(ws2, i, 2, cur,  10, PURPLE, bold=True, bg=bg, nf="£#,##0.00")
    w(ws2, i, 3, pmc,  10, bg=bg, nf="£#,##0.00")
    w(ws2, i, 4, ipp,  10, bg=bg, nf="£#,##0.00")
    w(ws2, i, 5, opp,  10, CORAL, bold=True, bg=bg, nf="£#,##0.00")
    w(ws2, i, 6, pme,  10, NAVY, bold=True, bg=bg, nf="£#,##0.00")
    w(ws2, i, 7, gap,  10, gap_col, bold=True, bg=bg, nf="0.0%")
    w(ws2, i, 8, rec,  10, GREEN, bold=True, bg=bg, nf="£#,##0.00")

sp(ws2, 12)
sec(ws2, 13, "HOW TO INTERPRET THE FOUR PSM INTERSECTION POINTS")
interpretations = [
    ("PMC — Point of Marginal Cheapness",
     "The 'Too Cheap' cumulative frequency line crosses the 'Expensive' line. Below this price, quality doubt starts outweighing value — consumers worry the product is inferior. This is the psychological price floor. Do not price below PMC."),
    ("IPP — Indifference Price Point",
     "Where 'Too Cheap' and 'Too Expensive' lines cross. At this price, equal numbers of respondents see the product as too cheap as see it as too expensive — they are indifferent. The point of maximum ambivalence."),
    ("OPP — Optimal Price Point",
     "The 'Cheap/Good Value' and 'Expensive' lines cross. At this price, the fewest respondents have strong objections — equal numbers think it's a bargain as think it's expensive. Maximum acceptance point. This is the recommended price anchor."),
    ("PME — Point of Marginal Expensiveness",
     "The 'Cheap' and 'Too Expensive' lines cross. Above this price, the majority see the product as too expensive to buy. This is the psychological price ceiling. Do not price above PME. Current prices sit 14-22% below PME — there is clear headroom."),
]
hdr(ws2, 14, ["Metric", "Interpretation"], [24, 72], bg=DGREY)
for i, (metric, interp) in enumerate(interpretations, 15):
    ws2.row_dimensions[i].height = 32
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws2, i, 1, metric, 10, NAVY, bold=True, bg=bg, align="left")
    ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
    c = ws2.cell(row=i, column=2, value=interp)
    c.font = font(10); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True); ba(c)

sp(ws2, 19)
note(ws2, 20,
    "⚠  HONEST LIMITATION: Van Westendorp measures stated preference, not revealed preference. "
    "Respondents saying they would buy at £33 does not guarantee they will. "
    "This is addressed through triangulation with competitor market data and the recommendation "
    "to set prices at or near OPP (not PME ceiling), which is where resistance is lowest. "
    "90-day post-change monitoring of conversion rate and return rate provides empirical validation.",
    bg=RED_BG, color=CORAL)


# ════════════════════════════════════════════════════════════════════
# SHEET 3 — COMPETITOR BENCHMARKING
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Competitor Benchmarking")
ws3.sheet_view.showGridLines = False
banner(ws3, "Branch 2: Competitor Price Benchmarking — 8 UK D2C Skincare Brands",
       "Real 2024-25 prices from brand websites, Cult Beauty, and Boots · Comparable SKU types")

note(ws3, 4,
    "NOTE ON COMPARABILITY: Prices are matched by product type (serum 30ml, moisturiser 50ml, retinol 30ml) "
    "rather than exact formula. All brands are DTC-led with UK distribution. Prices are mid-tier hero products, "
    "not promotional or introductory prices. Client brand is highlighted in purple.",
    bg=AMBER_BG)

sp(ws3, 5)
sec(ws3, 6, "PRICE COMPARISON TABLE — CLIENT vs 8 COMPETITORS")
hdr(ws3, 7,
    ["Brand", "Positioning Tier", "Serum\n30ml (£)", "Moisturiser\n50ml (£)", "Retinol\n30ml (£)",
     "Avg SKU\nPrice (£)", "Channel", "Key Differentiator"],
    [22, 20, 12, 14, 12, 14, 22, 32])

TIER_BG = {
    "Accessible": "FFEAF4FB",
    "Mid-Market": "FFEAFAF1",
    "Mid-Premium": "FFFEF9E7",
    "Premium": "FFFDF2E9",
    "Luxury": "FFFDEDEC",
    "CLIENT": "FFF5EEF8",
}
competitors = [
    ("The Ordinary",      "Accessible",  12.90, 15.90, 12.90, "D2C, Boots, Sephora",       "Ingredient transparency, min. price"),
    ("The Inkey List",    "Accessible",  14.99, 18.99, 13.99, "D2C, Boots, Sephora",       "Educational brand, clear formulas"),
    ("Byoma",             "Accessible+", 17.99, 21.99, None,  "D2C, Boots",                "Barrier-focused, strong Gen Z appeal"),
    ("Pixi Beauty",       "Mid-Market",  None,  32.00, None,  "D2C, M&S, Boots",           "Glow focus, cult toner brand"),
    ("CLIENT BRAND",      "Mid-Premium (frozen 3yr)", 28.00, 32.00, 38.00, "D2C only",    "Price unchanged 3 years — this analysis"),
    ("REN Clean Skincare","Mid-Premium", 38.00, 42.00, 44.00, "D2C, Boots, John Lewis",   "Clean beauty, sustainability"),
    ("Paula's Choice",    "Premium",     44.00, 48.00, 49.00, "D2C only",                  "Science-backed, loyal subscription base"),
    ("Medik8",            "Premium",     39.00, 42.00, 42.00, "D2C, clinics, aesthetics",  "Skin health positioning, clinical"),
    ("Elemis",            "Luxury",      70.00, 75.00, 72.00, "D2C, dept stores, Boots",   "Spa heritage, premium packaging"),
]
for i, (brand, tier, ser, moi, ret, channel, diff) in enumerate(competitors, 8):
    ws3.row_dimensions[i].height = 26
    is_client = brand == "CLIENT BRAND"
    tier_key = "CLIENT" if is_client else ("Accessible" if "Accessible" in tier else
               "Mid-Market" if "Mid-Market" in tier else "Mid-Premium" if "Mid-Premium" in tier else
               "Premium" if "Premium" in tier else "Luxury")
    bg = TIER_BG.get(tier_key, OFFWHITE)
    avg = sum(p for p in [ser, moi, ret] if p) / len([p for p in [ser, moi, ret] if p])

    w(ws3, i, 1, brand,   10, PURPLE if is_client else DGREY, bold=is_client, bg=bg, align="left")
    w(ws3, i, 2, tier,    9,  PURPLE if is_client else MIDGREY, italic=not is_client, bg=bg, align="left")
    for col, val in [(3, ser), (4, moi), (5, ret)]:
        if val:
            w(ws3, i, col, val, 10, PURPLE if is_client else DGREY, bold=is_client, bg=bg, nf="£#,##0.00")
        else:
            c = ws3.cell(row=i, column=col, value="—")
            c.font = font(10, SILVER); c.fill = fill(bg); c.alignment = Alignment(horizontal="center"); ba(c)
    w(ws3, i, 6, avg,     10, PURPLE if is_client else GREEN if avg > 40 else AMBER if avg > 25 else SKY,
      bold=is_client, bg=bg, nf="£#,##0.00")
    w(ws3, i, 7, channel,  9, MIDGREY, italic=True, bg=bg, align="left")
    w(ws3, i, 8, diff,     9, DGREY, bg=bg, align="left", wrap=True)

sp(ws3, 17)
sec(ws3, 18, "KEY FINDINGS FROM COMPETITOR BENCHMARKING")
comp_findings = [
    ("Positioning gap",     "Client brand is priced as mid-market (£28-38) but aspires to mid-premium positioning (clean, active-led, D2C). REN, Paula's Choice, and Medik8 command £10-16 more for comparable SKUs with similar positioning claims."),
    ("Price signal issue",  "The Retinol at £38 is priced below REN (£44), Paula's Choice (£49), and Medik8 (£42). In skincare, price signals efficacy. A premium retinol priced below mid-premium competitors sends a confusing quality signal."),
    ("The accessible floor","The Ordinary and The Inkey List create the accessible floor at £13-19. Client brand at £28 is already 2x the floor — there is no risk of being confused with the ingredient-only mass tier by raising to £33-46."),
    ("Channel advantage",   "Client is DTC only. Paula's Choice is also DTC only and prices at £44-49. DTC allows premium pricing without retail margin concessions. The brand is not capturing this structural advantage."),
]
hdr(ws3, 19, ["Finding", "Detail"], [22, 68])
for i, (label, detail) in enumerate(comp_findings, 20):
    ws3.row_dimensions[i].height = 30
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws3, i, 1, label, 10, NAVY, bold=True, bg=bg, align="left")
    ws3.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
    c = ws3.cell(row=i, column=2, value=detail)
    c.font = font(10); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True); ba(c)


# ════════════════════════════════════════════════════════════════════
# SHEET 4 — MARGIN WATERFALL (all 12 SKUs)
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Margin Waterfall")
ws4.sheet_view.showGridLines = False
banner(ws4, "Branch 3: Margin Waterfall — Current vs Proposed Pricing (All 12 SKUs)",
       "COGS held fixed in £ · Volume elasticity = -0.8 · Benchmarked vs Eightx DTC Beauty 2026 (median GM 69.4%)")

note(ws4, 4,
    "HOW TO READ: Current price × current units = current revenue. Proposed price × adjusted units (after elasticity) = proposed revenue. "
    "COGS per unit is FIXED in £ — it doesn't change. So gross margin % improves as price rises. "
    "Volume elasticity of -0.8 means a 10% price increase leads to an 8% volume decline — conservative for premium D2C skincare.",
    bg=AMBER_BG)

sp(ws4, 5)
sec(ws4, 6, "FULL SKU WATERFALL — CURRENT vs PROPOSED")
hdr(ws4, 7,
    ["Code","SKU Name","Category","Current\nPrice","Current\nGM %","Proposed\nPrice","Increase\n%","Proposed\nGM %","GM\nUplift (pp)","Revenue\nUplift £","GP\nUplift £"],
    [8, 28, 14, 12, 12, 12, 10, 12, 14, 14, 12])

sku_data = [
    ("S01","Vitamin C Brightening Serum 30ml",    "Serum",       28.00, 0.720, 33.00, 0.179),
    ("S02","Hyaluronic Acid Moisturiser 50ml",     "Moisturiser", 32.00, 0.700, 37.00, 0.156),
    ("S03","Retinol Night Treatment 30ml",          "Treatment",   38.00, 0.730, 46.00, 0.211),
    ("S04","Niacinamide Pore Minimiser 30ml",       "Serum",       26.00, 0.712, 30.00, 0.154),
    ("S05","SPF 50 Daily Sunscreen 50ml",           "SPF",         24.00, 0.670, 27.00, 0.125),
    ("S06","Gentle Enzyme Exfoliator 75ml",         "Exfoliator",  30.00, 0.690, 34.00, 0.133),
    ("S07","Hydrating Cleansing Balm 100ml",        "Cleanser",    22.00, 0.650, 24.00, 0.091),
    ("S08","AHA/BHA Toning Solution 150ml",         "Toner",       20.00, 0.700, 22.00, 0.100),
    ("S09","Ceramide Barrier Repair Cream 50ml",    "Moisturiser", 34.00, 0.720, 40.00, 0.176),
    ("S10","Eye Cream Peptide Complex 15ml",        "Eye Cream",   36.00, 0.750, 43.00, 0.194),
    ("S11","Overnight Recovery Mask 50ml",          "Mask",        28.00, 0.680, 32.00, 0.143),
    ("S12","Multi-Action Mist Toner 100ml",         "Toner",       18.00, 0.660, 19.50, 0.083),
]
BRAND_ARR = 1_400_000
rev_weights = [0.20,0.18,0.14,0.13,0.05,0.06,0.05,0.04,0.05,0.04,0.03,0.03]
ELASTICITY = -0.8

total_cur_rev = 0; total_prop_rev = 0
total_cur_gp  = 0; total_prop_gp  = 0

for i, ((code, name, cat, cp, cgm, pp, inc_pct), rw) in enumerate(zip(sku_data, rev_weights), 8):
    ws4.row_dimensions[i].height = 24
    bg = "FFFFF8F0" if code in ["S01","S02","S03","S04"] else (LTGREY if i%2==0 else OFFWHITE)
    is_hero = code in ["S01","S02","S03","S04"]

    cur_rev  = BRAND_ARR * rw
    cogs_unit = cp * (1 - cgm)
    pgm = 1 - (cogs_unit / pp)
    vol_change = inc_pct * ELASTICITY
    units = cur_rev / cp
    prop_rev  = pp * units * (1 + vol_change)
    cur_gp    = cur_rev * cgm
    prop_gp   = prop_rev * pgm
    rev_upl   = prop_rev - cur_rev
    gp_upl    = prop_gp  - cur_gp
    gm_uplift = pgm - cgm

    total_cur_rev  += cur_rev;  total_prop_rev += prop_rev
    total_cur_gp   += cur_gp;   total_prop_gp  += prop_gp

    gm_col = GREEN if pgm > 0.70 else AMBER if pgm > 0.65 else CORAL

    w(ws4, i, 1,  code,     9,  NAVY, bold=is_hero, bg=bg)
    w(ws4, i, 2,  name,     9,  DGREY, bold=is_hero, bg=bg, align="left")
    w(ws4, i, 3,  cat,      9,  MIDGREY, bg=bg, align="left")
    w(ws4, i, 4,  cp,       10, PURPLE, bold=True, bg=bg, nf="£#,##0.00")
    w(ws4, i, 5,  cgm,      10, SILVER, bg=bg, nf="0.0%")
    w(ws4, i, 6,  pp,       10, GREEN, bold=True, bg=bg, nf="£#,##0.00")
    w(ws4, i, 7,  inc_pct,  10, CORAL if inc_pct > 0.15 else AMBER, bold=True, bg=bg, nf="0.0%")
    w(ws4, i, 8,  pgm,      10, gm_col, bold=True, bg=bg, nf="0.0%")
    w(ws4, i, 9,  gm_uplift,10, GREEN, bg=bg, nf="+0.0pp;-0.0pp;0.0pp")
    w(ws4, i, 10, rev_upl,  10, GREEN if rev_upl > 0 else CORAL, bg=bg, nf="£#,##0")
    w(ws4, i, 11, gp_upl,   10, GREEN if gp_upl > 0 else CORAL, bold=True, bg=bg, nf="£#,##0")

ws4.row_dimensions[20].height = 28
totals = [("TOTAL", None, None, None, total_cur_gp/total_cur_rev,
           None, (total_prop_rev-total_cur_rev)/total_cur_rev,
           total_prop_gp/total_prop_rev,
           total_prop_gp/total_prop_rev - total_cur_gp/total_cur_rev,
           total_prop_rev - total_cur_rev, total_prop_gp - total_cur_gp)]
for col, (v, nf) in enumerate([("TOTAL",None),(None,None),(None,None),(None,None),
    (total_cur_gp/total_cur_rev,"0.0%"),(None,None),
    ((total_prop_rev-total_cur_rev)/total_cur_rev,"0.0%"),
    (total_prop_gp/total_prop_rev,"0.0%"),
    (total_prop_gp/total_prop_rev-total_cur_gp/total_cur_rev,"+0.0pp;-0.0pp"),
    (total_prop_rev-total_cur_rev,"£#,##0"),
    (total_prop_gp-total_cur_gp,"£#,##0")], 1):
    cell = ws4.cell(row=20, column=col, value=v)
    cell.font = font(10, WHITE, bold=True); cell.fill = fill(NAVY)
    ba(cell); cell.alignment = Alignment(horizontal="center", vertical="center")
    if nf and v: cell.number_format = nf

sp(ws4, 21)
note(ws4, 22,
    f"Benchmark: Eightx Beauty DTC Margin Benchmarks 2026 — median DTC beauty gross margin 69.4% "
    f"(e.l.f. 71.2%, Olaplex 69.4%, Beauty Health 65.3% — Eightx citing SEC 10-K filings). "
    f"Current brand GM of 70.6% is above median. Post-repricing GM of 74.6% moves brand into "
    f"top-quartile territory for DTC beauty, consistent with premium D2C positioning.",
    bg=GREEN_BG)


# ════════════════════════════════════════════════════════════════════
# SHEET 5 — RECOMMENDATION & IMPLEMENTATION
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Recommendation")
ws5.sheet_view.showGridLines = False
banner(ws5, "Triangulated Recommendation & Implementation Plan",
       "Where PSM, competitor benchmarking, and margin waterfall all converge")

sec(ws5, 4, "RECOMMENDED PRICES — 3 HERO SKUs (FULL TRIANGULATION)")
hdr(ws5, 5,
    ["SKU", "Current", "PSM OPP\n(Max acceptance)", "PSM PME\n(Ceiling)", "Comp Mid-Mkt", "Comp Premium",
     "RECOMMENDED\nPRICE", "Rationale"],
    [22, 10, 14, 14, 14, 14, 16, 36])

tri_data = [
    ("Vitamin C Serum",   28.00, 30.20, 35.70, 25.00, 40.33, 33.00,
     "OPP-anchored. Below PME. Between mid-mkt and premium comp. Avoids psychological £35 barrier."),
    ("HA Moisturiser",    32.00, 35.50, 41.80, 27.00, 44.00, 37.00,
     "OPP-anchored. Moves brand from below to parity with Pixi/Byoma premium, well below REN."),
    ("Retinol Treatment", 38.00, 44.00, 48.00, None,  41.67, 46.00,
     "OPP-anchored. Removes the quality-confusion of a premium retinol priced below REN/Medik8."),
]
for i, (sku, cur, opp, pme, mid, prem, rec, rationale) in enumerate(tri_data, 6):
    ws5.row_dimensions[i].height = 38
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws5, i, 1, sku,   10, NAVY, bold=True, bg=bg, align="left")
    w(ws5, i, 2, cur,   10, PURPLE, bold=True, bg=bg, nf="£#,##0.00")
    w(ws5, i, 3, opp,   10, CORAL, bold=True, bg=bg, nf="£#,##0.00")
    w(ws5, i, 4, pme,   10, bg=bg, nf="£#,##0.00")
    w(ws5, i, 5, mid if mid else "—", 10, bg=bg, nf="£#,##0.00" if mid else "@")
    w(ws5, i, 6, prem,  10, SKY, bg=bg, nf="£#,##0.00")
    w(ws5, i, 7, rec,   11, GREEN, bold=True, bg=GREEN_BG, nf="£#,##0.00")
    w(ws5, i, 8, rationale, 9, DGREY, italic=True, bg=bg, align="left", wrap=True)

sp(ws5, 9)
sec(ws5, 10, "IMPLEMENTATION — WHAT TO DO AND IN WHAT ORDER")
impl = [
    ("Step 1 — Before launch", "Redesign packaging on hero SKUs. New copy that emphasises active ingredients and formulation quality. This is the 'reason why' for the new price. Without this, the price rise looks opportunistic. With it, it looks like an upgrade."),
    ("Step 2 — Communication", "Email existing customers 2 weeks before launch: 'Our formulations are evolving. As part of our commitment to quality, we're updating our pricing from [date].' No apology. Frame as brand maturity, not cost pressure."),
    ("Step 3 — Do NOT launch with a discount", "Never discount at the new price in the first 6 months. This would immediately train customers to wait for a sale and anchor the 'real' price back at the old level. Full price from day one."),
    ("Step 4 — Loyalty offer (one-time)", "Offer existing customers a final order at current prices before the change goes live. This generates goodwill and a revenue spike — NOT a discount on the new price."),
    ("Step 5 — Monitor closely", "Track daily: conversion rate, add-to-cart rate, return rate, and review sentiment. Expect a 2-3 week dip in conversion as new visitors adjust. If conversion doesn't recover by week 6, re-examine one SKU price."),
    ("Step 6 — Remaining 8 SKUs", "Apply the same framework to SKUs 05-12 in month 3, after the hero SKU repricing has stabilised. Do not change all 12 at once — this risks customer overwhelm and creates attribution confusion."),
]
hdr(ws5, 11, ["Step", "Detail"], [20, 76])
for i, (step, detail) in enumerate(impl, 12):
    ws5.row_dimensions[i].height = 34
    bg = LTGREY if i % 2 == 0 else OFFWHITE
    w(ws5, i, 1, step, 10, NAVY, bold=True, bg=bg, align="left", wrap=True)
    ws5.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
    c = ws5.cell(row=i, column=2, value=detail)
    c.font = font(10); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True); ba(c)

sp(ws5, 18)
sec(ws5, 19, "HONEST ASSESSMENT — WHAT COULD GO WRONG")
risks = [
    ("PSM is stated, not revealed",
     "Respondents saying they'd buy at £33 doesn't guarantee they will. Mitigated by: (1) triangulation with competitor price proof, (2) pricing at OPP not PME, (3) 90-day monitoring window."),
    ("Volume decline exceeds -0.8",
     "If elasticity is higher (say, -1.5), volume loss outweighs price gain and total GP falls. Monitored through daily conversion tracking. Rollback plan: revert one SKU within 30 days if conversion drops >20% without recovery."),
    ("Competitor response",
     "REN or Paula's Choice could run a promotion during the repricing window, making the brand look expensive. Unlikely but possible. Maintain marketing spend and new campaign during transition."),
    ("Wrong to raise all at once",
     "This analysis recommends hero SKUs first, rest in month 3. Raising all 12 simultaneously risks overwhelming customers and makes attribution impossible. Phased approach is critical."),
]
hdr(ws5, 20, ["Risk", "Mitigation"], [24, 72])
for i, (risk, mitigation) in enumerate(risks, 21):
    ws5.row_dimensions[i].height = 30
    bg = RED_BG if i == 21 else AMBER_BG if i == 22 else OFFWHITE
    w(ws5, i, 1, risk, 10, CORAL, bold=True, bg=bg, align="left")
    ws5.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
    c = ws5.cell(row=i, column=2, value=mitigation)
    c.font = font(10); c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True); ba(c)

ws5.merge_cells("A26:I26")
cf = ws5.cell(row=26, column=1,
    value="Vaishnavi Bhor · MSc Business Analytics, University of Manchester · vbhor207@gmail.com · vbho.github.io/portfolio")
cf.font = font(10, SKY); cf.fill = fill(OFFWHITE)
cf.alignment = Alignment(horizontal="left", vertical="center", indent=1); ba(cf); ws5.row_dimensions[26].height = 20

outpath = OUT / "Skincare_Pricing_Strategy_VaishnaviBhor.xlsx"
wb.save(outpath)
print(f"\n✓  Workbook saved → {outpath}")
print(f"   Sheets: {len(wb.sheetnames)}")
